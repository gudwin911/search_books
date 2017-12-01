import xlwt
import xlsxwriter
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import Page
import webbrowser
from selenium import webdriver
import re
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook

shop = "https://lavkababuin.com/"
# field = "keyword"
book_list = ["Хрома", "Nobrow. Культура маркетинга. Маркетинг культуры", "123", "ДНК особистості"]


def create_file(shop):
    wb = xlsxwriter.Workbook("price_analysis.xlsx")
    ws = wb.add_worksheet("prices")
    bold = wb.add_format({'bold': True})
    ws.write("A1", "Книга", bold)
    ws.write("B1", shop, bold)
    wb.close()


# create_file(book)


def get_book():
    # webbrowser.open("https://lavkababuin.com/")
    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    driver.get(shop)
    lst = []
    for item in book_list:
        search = driver.find_element(By.ID, "keyword")
        # WebDriverWait(driver, 10).until(
        #     EC.element_to_be_clickable((By.ID, "keyword")))
        search.send_keys(item)
        search.submit()
        # WebDriverWait(driver, 10).until(
        #     EC.visibility_of_element_located((By.TAG_NAME, "h1")))
        try:
            book = driver.find_element(By.LINK_TEXT, item)
            book.click()
            price = driver.find_element(By.ID, "yproduct_price").text
            lst.append(price)
        except NoSuchElementException:
            driver.find_element(By.ID, "keyword").clear()
            lst.append("0")
        #WebDriverWait(driver, 10).until(
        #    EC.visibility_of_element_located((By.ID, "yproduct_price")))
        # price = driver.find_element(By.ID, "yproduct_price").text
        # lst.append(price)
    driver.close()
    return lst

# get_book()

# a = ['378 грн.', '315 грн.']
def make_price_without_literals():
    a = get_book()
    # to delete literals from price e.g "300 $" => "300"
    lst =[]
    for i in a:
        reg = re.compile("[^0-9]")
        i = reg.sub("", i)
        lst.append(i)
    return lst

# make_price_without_literals()


def add_prices_to_list():
    # lst = make_price_without_literals()
    lst = ["375", "315", "0", "192"]
    wb = xlsxwriter.Workbook("price_analysis.xlsx")
    ws = wb.add_worksheet("prices")
    row = 1
    col = 0
    num = 1
    for book, price in zip(book_list, lst):
        ws.write(row, col, num)
        ws.write(row, col + 1, str(book))
        ws.write(row, col + 2, int(price))
        row += 1
        num += 1
    wb.close()

add_prices_to_list()

# from datetime import datetime
#
# style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
#     num_format_str='#,##0.00')
# style1 = xlwt.easyxf(num_format_str='D-MMM-YY')
#
# wb = xlwt.Workbook()
# ws = wb.add_sheet('A Test Sheet')
#
# ws.write(0, 0, 1234.56, style0)
# ws.write(1, 0, datetime.now(), style1)
# ws.write(2, 0, 1)
# ws.write(2, 1, 1)
# ws.write(2, 2, xlwt.Formula("A3+B3"))
#
# wb.save('example.xls')
#
# #######################################################
#
# # Create an new Excel file and add a worksheet.
# workbook = xlsxwriter.Workbook('demo.xlsx')
# worksheet = workbook.add_worksheet()
#
# # Widen the first column to make the text clearer.
# worksheet.set_column('A:A', 20)
#
# # Add a bold format to use to highlight cells.
# bold = workbook.add_format({'bold': True})
#
# # Write some simple text.
# worksheet.write('A1', 'Hello')
#
# # Text with formatting.
# worksheet.write('A2', 'World', bold)
#
# # Write some numbers, with row/column notation.
# worksheet.write(2, 0, 123)
# worksheet.write(3, 0, 123.456)
#
# # Insert an image.
# worksheet.insert_image('B5', 'logo.png')
#
# workbook.close()